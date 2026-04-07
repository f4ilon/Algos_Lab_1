from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker


filename = "filet.xlsx"
wb = load_workbook(filename)

def save_results(results, sheetname):
    if sheetname in wb.sheetnames:
        del wb[sheetname]
    ws = wb.create_sheet(sheetname)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25

    # Записываем данные
    ws.append(["size", "time(ms)"])
    for row, (int_val, float_val) in enumerate(results, start=2):
        cell_a = ws.cell(row=row, column=1, value=int_val)
        cell_a.number_format = '0'

        cell_b = ws.cell(row=row, column=2, value=float_val)
        cell_b.number_format = '0.0000000000'

    # Создаём диаграмму
    chart = ScatterChart()
    chart.title = "График зависимости Y от X"
    chart.x_axis.title = "size"
    chart.y_axis.title = "time"
    chart.style = 13

    x_refs = Reference(ws, min_col=1, min_row=2, max_row=len(results)-1)
    y_refs = Reference(ws, min_col=2, min_row=2, max_row=len(results)-1)

    series = Series(y_refs, x_refs, title="Серия 1")

    series.smooth = False
    series.graphicalProperties.line.solidFill = "0000FF"
    series.graphicalProperties.line.prstDash = "solid"
    series.graphicalProperties.line.width = 25000
    series.marker = Marker('circle')

    chart.series.append(series)

    ws.add_chart(chart, "D2")
    wb.save(filename)


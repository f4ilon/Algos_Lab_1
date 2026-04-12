from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
import os


script_path = os.path.abspath(__file__)
script_dir = os.path.dirname(script_path)
parent_dir = os.path.dirname(script_dir) 

filename = os.path.join(parent_dir, 'file.xlsx')
wb = load_workbook(filename)

def save_results(results, sheetname):
    if sheetname in wb.sheetnames:
        del wb[sheetname]
    ws = wb.create_sheet(sheetname)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25

    # Записываем данные
    ws.append(["n", "time(ms)", "time/n"])
    for row, (size, time) in enumerate(results, start=2):
        cell_a = ws.cell(row=row, column=1, value=size)
        cell_a.number_format = '0'

        cell_b = ws.cell(row=row, column=2, value=time)
        cell_b.number_format = '0.0000000000'

        cell_c = ws.cell(row=row, column=3, value=time/size)
        cell_c.number_format = '0.0000000000'

    # Создаём диаграмму
    chart = ScatterChart()
    chart.x_axis.title = "n"
    chart.y_axis.title = "time(ms)"
    chart.style = 13

    x_refs = Reference(ws, min_col=1, min_row=2, max_row=len(results)-1)
    y_refs = Reference(ws, min_col=2, min_row=2, max_row=len(results)-1)

    series = Series(y_refs, x_refs, title="time(n)")

    series.smooth = False
    series.graphicalProperties.line.solidFill = "0000FF"
    series.graphicalProperties.line.prstDash = "solid"
    series.graphicalProperties.line.width = 25000
    series.marker = Marker('circle')

    chart.series.append(series)

    ws.add_chart(chart, "D2")

    chart = ScatterChart()
    chart.x_axis.title = "n"
    chart.y_axis.title = "y"
    chart.style = 13

    x_refs = Reference(ws, min_col=1, min_row=2, max_row=len(results)-1)
    y_refs = Reference(ws, min_col=3, min_row=2, max_row=len(results)-1)

    series = Series(y_refs, x_refs, title="y=time/n")

    series.smooth = False
    series.graphicalProperties.line.solidFill = "0000FF"
    series.graphicalProperties.line.prstDash = "solid"
    series.graphicalProperties.line.width = 25000
    series.marker = Marker('circle')

    chart.series.append(series)

    ws.add_chart(chart, "M2")
    wb.save(filename)

